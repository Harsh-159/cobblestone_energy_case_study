"""
Daily report generation for DE_LU power trading intelligence.

Generates:
  - Two matplotlib figures (signal dashboard + tomorrow's forecast)
  - Raw data Excel workbook (5 sheets)
  - PDF report (3 pages via WeasyPrint + Jinja2)

Reads only from saved pipeline output files — does not re-run any pipeline step.
"""
from __future__ import annotations

import json
import logging
import sys
from datetime import datetime
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.patches as mpatches
import numpy as np
import pandas as pd

# ── Project paths ────────────────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT_ROOT / "data"
OUTPUTS_DIR = PROJECT_ROOT / "outputs"
REPORTS_DIR = OUTPUTS_DIR / "reports"
TEMPLATES_DIR = PROJECT_ROOT / "server" / "templates"

# ── Plot style ───────────────────────────────────────────────────────────
BG_DARK = "#0f1117"
BG_CARD = "#1a1f2e"
TEXT_LIGHT = "#e2e8f0"
TEXT_MUTED = "#9ca3af"
BLUE = "#2563EB"
RED = "#DC2626"
GREEN = "#16A34A"
ORANGE = "#D97706"
PURPLE = "#7C3AED"
PLOT_DPI = 150

logger = logging.getLogger("report_generator")


def _setup_logging():
    if not logger.handlers:
        logger.setLevel(logging.DEBUG)
        fmt = logging.Formatter("%(asctime)s | %(levelname)-8s | %(message)s",
                                datefmt="%Y-%m-%d %H:%M:%S")
        ch = logging.StreamHandler(sys.stdout)
        ch.setLevel(logging.INFO)
        ch.setFormatter(fmt)
        logger.addHandler(ch)


def _dark_style():
    plt.rcParams.update({
        "figure.facecolor": BG_DARK,
        "axes.facecolor": BG_DARK,
        "axes.edgecolor": "#374151",
        "axes.labelcolor": TEXT_LIGHT,
        "text.color": TEXT_LIGHT,
        "xtick.color": TEXT_MUTED,
        "ytick.color": TEXT_MUTED,
        "axes.grid": True,
        "grid.color": "#1f2937",
        "grid.alpha": 0.5,
        "grid.linestyle": "--",
        "font.size": 10,
    })


# =====================================================================
# Figure 1: Signal Dashboard
# =====================================================================
def _generate_fig1(signal_df, delivery_df, target_date, out_dir):
    """Three-panel trading signal view."""
    _dark_style()

    td = pd.Timestamp(target_date)
    window_start = td - pd.Timedelta(days=30)
    # Only extend forward to where data actually exists (not empty future)
    max_date = signal_df["delivery_date"].max()
    window_end = min(td + pd.Timedelta(days=7), max_date)

    sig = signal_df[
        (signal_df["delivery_date"] >= window_start)
        & (signal_df["delivery_date"] <= window_end)
    ].copy()

    dates = sig["delivery_date"]

    fig = plt.figure(figsize=(14, 8))
    gs = fig.add_gridspec(2, 2, height_ratios=[1.5, 1], hspace=0.35, wspace=0.3)

    # ── Panel 1: Fair value vs actual vs curve proxy ──────────────────
    ax1 = fig.add_subplot(gs[0, :])

    if "base_actual" in sig.columns:
        ax1.plot(dates, sig["base_actual"], color="#6b7280", linewidth=0.8,
                 label="DA actual (base)", alpha=0.7)
    if "fv_month_base" in sig.columns:
        ax1.plot(dates, sig["fv_month_base"], color=BLUE, linewidth=1.2,
                 label="Fair value (M+1 base)")
    if "curve_proxy_30d" in sig.columns:
        ax1.plot(dates, sig["curve_proxy_30d"], color=ORANGE, linewidth=1.0,
                 linestyle="--", label="Curve proxy (30d)")
    if "fv_month_base_p10" in sig.columns and "fv_month_base_p90" in sig.columns:
        ax1.fill_between(dates, sig["fv_month_base_p10"], sig["fv_month_base_p90"],
                         alpha=0.12, color=BLUE, label="P10–P90 band")

    # Signal background strips
    for _, row in sig.iterrows():
        lbl = str(row.get("signal_label_month_base", ""))
        if "BUY" in lbl:
            c = GREEN
        elif "SELL" in lbl:
            c = RED
        elif "INVALIDATED" in lbl:
            c = "#fbbf24"
        else:
            c = "#374151"
        ax1.axvspan(row["delivery_date"] - pd.Timedelta(hours=12),
                    row["delivery_date"] + pd.Timedelta(hours=12),
                    alpha=0.06, color=c, zorder=0)

    # Today marker
    ax1.axvline(td, color="white", linestyle="--", linewidth=0.8, alpha=0.7)
    ax1.text(td, ax1.get_ylim()[1] * 0.98, " Today", color="white",
             fontsize=8, va="top")

    fwd_days = (window_end - td).days
    title_suffix = f" + {fwd_days}-day forward" if fwd_days > 0 else ""
    ax1.set_title(f"M+1 base fair value vs curve proxy — last 30 days{title_suffix}",
                  fontsize=11)
    ax1.set_ylabel("Price (€/MWh)")
    ax1.legend(fontsize=7, loc="upper left", framealpha=0.3)
    ax1.xaxis.set_major_formatter(mdates.DateFormatter("%m-%d"))
    ax1.xaxis.set_major_locator(mdates.WeekdayLocator(interval=1))
    plt.setp(ax1.xaxis.get_majorticklabels(), rotation=30, ha="right")

    # ── Panel 2: Signal premium bars ──────────────────────────────────
    ax2 = fig.add_subplot(gs[1, 0])
    hist = sig[sig["delivery_date"] <= td].tail(30)
    premium = hist["signal_month_base"].values
    hdates = hist["delivery_date"]

    bar_colors = [BLUE if (pd.notna(p) and p >= 0) else RED for p in premium]
    ax2.bar(hdates, np.where(np.isnan(premium), 0, premium),
            color=bar_colors, width=0.8, edgecolor="none")

    # Today's bar outline
    today_mask = hist["delivery_date"] == td
    if today_mask.any():
        today_val = hist.loc[today_mask, "signal_month_base"].iloc[0]
        if pd.notna(today_val):
            ax2.bar(td, today_val, width=0.8, color="none",
                    edgecolor="white", linewidth=1.5)
            ax2.annotate(f"{today_val:+.1f}", (td, today_val),
                         textcoords="offset points", xytext=(0, 8 if today_val >= 0 else -12),
                         fontsize=7, color="white", ha="center")

    for thr in [3, 8, -3, -8]:
        ax2.axhline(thr, color="#374151", linestyle="--", linewidth=0.6)
    ax2.axhline(0, color="#6b7280", linewidth=0.5)
    ax2.set_title("Daily signal premium (€/MWh)", fontsize=10)
    ax2.set_ylabel("Premium (€/MWh)")
    ax2.xaxis.set_major_formatter(mdates.DateFormatter("%m-%d"))
    plt.setp(ax2.xaxis.get_majorticklabels(), rotation=30, ha="right", fontsize=7)

    # ── Panel 3: Today's scorecard ────────────────────────────────────
    ax3 = fig.add_subplot(gs[1, 1])
    ax3.set_xlim(0, 10)
    ax3.set_ylim(0, 10)
    ax3.axis("off")

    today_row = sig[sig["delivery_date"] == td]
    if len(today_row) > 0:
        r = today_row.iloc[0]
        sig_label_raw = str(r.get("signal_label_month_base", "N/A"))
        # Normalise non-standard labels to HOLD
        sig_label = sig_label_raw if sig_label_raw in (
            "STRONG_BUY", "BUY", "HOLD", "SELL", "STRONG_SELL", "INVALIDATED"
        ) else "HOLD"
        sig_color = GREEN if "BUY" in sig_label else RED if "SELL" in sig_label else "#fbbf24" if "INV" in sig_label else TEXT_MUTED

        def _txt(y, left, right, color=TEXT_LIGHT):
            ax3.text(0.3, y, left, fontsize=7.5, color=TEXT_MUTED, va="center",
                     fontfamily="monospace")
            ax3.text(6.5, y, str(right), fontsize=7.5, color=color, va="center",
                     fontfamily="monospace", fontweight="bold")

        arrow = "▲" if "BUY" in sig_label else "▼" if "SELL" in sig_label else "●"
        ax3.text(0.3, 9.5, "TODAY'S SIGNAL", fontsize=8, color=BLUE, fontweight="bold")
        ax3.text(0.3, 9.0, "─" * 20, fontsize=6, color="#374151")
        _txt(8.4, "Direction:", f"{sig_label.replace('_', ' ')}  {arrow}", sig_color)
        _txt(7.8, "Product:", "M+1 Base")
        # Only show premium/confidence if they have actual values
        y_pos = 7.2
        if pd.notna(r.get("signal_month_base")):
            _txt(y_pos, "Premium:", f"{r.get('signal_month_base'):+.1f} €/MWh")
            y_pos -= 0.6
        if pd.notna(r.get("month_confidence")):
            _txt(y_pos, "Confidence:", f"{r.get('month_confidence'):.2f}")
            y_pos -= 0.6
        _txt(y_pos, "Position:", f"{r.get('month_position_size', 0):.1f}×")

        ax3.text(0.3, 5.2, "INVALIDATION", fontsize=8, color=BLUE, fontweight="bold")
        ax3.text(0.3, 4.8, "─" * 20, fontsize=6, color="#374151")
        inv_flags = [
            ("Gas spike:", "inv_gas_spike"),
            ("Wind rev.:", "inv_wind_revision"),
            ("Res. load:", "inv_residual_load_swing"),
            ("Neg. prices:", "inv_negative_price_regime"),
            ("High error:", "inv_high_error_regime"),
        ]
        for i, (label, key) in enumerate(inv_flags):
            val = r.get(key, False)
            icon = "✓" if not val else "✗"
            color = GREEN if not val else RED
            _txt(4.2 - i * 0.5, label, icon, color)

        inv_any = r.get("any_invalidation", False)
        status = "⚠ TRIGGERED" if inv_any else "CLEAN ✓"
        s_color = RED if inv_any else GREEN
        _txt(1.5, "Status:", status, s_color)

        # Spark spread
        spark = r.get("spark_spread_actual")
        if pd.notna(spark):
            ax3.text(6.5, 5.2, "SPARK SPREAD", fontsize=7, color=BLUE, fontweight="bold")
            ax3.text(6.5, 4.7, f"{spark:+.1f} €/MWh", fontsize=7.5,
                     color=GREEN if spark > 0 else RED, fontweight="bold")

    fig.savefig(out_dir / "fig_signal_dashboard.png", dpi=PLOT_DPI,
                bbox_inches="tight", facecolor=BG_DARK)
    plt.close(fig)
    logger.info("Saved fig_signal_dashboard.png")


# =====================================================================
# Figure 2: Tomorrow's Hourly Forecast
# =====================================================================
def _generate_fig2(oos_df, delivery_df, raw_df, target_date, out_dir):
    """Tomorrow's hourly price profile + seasonal comparison."""
    _dark_style()

    td = pd.Timestamp(target_date, tz="UTC")
    # "Tomorrow" = next delivery date after target
    tomorrow = td + pd.Timedelta(days=1)
    tomorrow_str = str(tomorrow.date())

    # Get hourly predictions for tomorrow
    tomorrow_data = oos_df.loc[
        (oos_df.index >= tomorrow)
        & (oos_df.index < tomorrow + pd.Timedelta(days=1))
    ].copy()

    if len(tomorrow_data) == 0:
        # Fall back to target date itself
        tomorrow_data = oos_df.loc[
            (oos_df.index >= td)
            & (oos_df.index < td + pd.Timedelta(days=1))
        ].copy()
        tomorrow_str = target_date

    if len(tomorrow_data) == 0:
        logger.warning("No hourly data for tomorrow — skipping fig2")
        return

    tomorrow_data["hour"] = tomorrow_data.index.hour

    # Get delivery period summary
    del_row = delivery_df[delivery_df["delivery_date"] == pd.Timestamp(tomorrow_str)]
    base_pred = del_row["base_pred"].iloc[0] if len(del_row) > 0 else tomorrow_data["y_pred"].mean()
    peak_pred = del_row["peak_pred"].iloc[0] if len(del_row) > 0 else np.nan
    peak_prem = del_row["peak_base_premium_pred"].iloc[0] if len(del_row) > 0 and "peak_base_premium_pred" in del_row.columns else np.nan

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 6),
                                    gridspec_kw={"width_ratios": [1.85, 1]})

    # ── Panel 1: Hourly forecast ──────────────────────────────────────
    hours = tomorrow_data["hour"].values
    preds = tomorrow_data["y_pred"].values

    ax1.plot(hours, preds, color=BLUE, linewidth=1.5, marker="o", markersize=4,
             label="Forecast", zorder=5)

    # Base and peak lines
    ax1.axhline(base_pred, color=ORANGE, linestyle="--", linewidth=0.8,
                label=f"Base avg: {base_pred:.1f}")
    if pd.notna(peak_pred):
        ax1.axhline(peak_pred, color=GREEN, linestyle="--", linewidth=0.8,
                    label=f"Peak avg: {peak_pred:.1f}")

    # Peak block shading (8-20 UTC)
    ax1.axvspan(8, 19, alpha=0.06, color=ORANGE, zorder=0)
    ax1.text(13.5, ax1.get_ylim()[0] if ax1.get_ylim()[0] != 0 else preds.min() * 0.95,
             "Peak block", fontsize=7, color=ORANGE, ha="center", alpha=0.7)

    # Min/max annotations
    if len(preds) > 0:
        min_idx = np.argmin(preds)
        max_idx = np.argmax(preds)
        ax1.scatter(hours[min_idx], preds[min_idx], color=RED, s=60, zorder=6)
        ax1.annotate(f"Min: {preds[min_idx]:.1f}\nH{hours[min_idx]:02d}",
                     (hours[min_idx], preds[min_idx]),
                     textcoords="offset points", xytext=(-15, -20),
                     fontsize=7, color=RED)
        ax1.scatter(hours[max_idx], preds[max_idx], color=GREEN, s=60, zorder=6)
        ax1.annotate(f"Max: {preds[max_idx]:.1f}\nH{hours[max_idx]:02d}",
                     (hours[max_idx], preds[max_idx]),
                     textcoords="offset points", xytext=(8, 8),
                     fontsize=7, color=GREEN)

    ax1.set_title(f"Tomorrow's hourly forecast — {tomorrow_str}", fontsize=11)
    ax1.set_xlabel("Hour (UTC)")
    ax1.set_ylabel("Forecast price (€/MWh)")
    ax1.set_xticks(range(0, 24))
    ax1.set_xticklabels([f"{h:02d}" for h in range(24)], fontsize=7)
    ax1.legend(fontsize=7, loc="upper left", framealpha=0.3)

    # Info box
    info = f"Base: {base_pred:.1f} €/MWh"
    if pd.notna(peak_pred):
        info += f" | Peak: {peak_pred:.1f} €/MWh"
    if pd.notna(peak_prem):
        info += f" | Prem: {peak_prem:+.1f}"
    props = dict(boxstyle="round,pad=0.4", facecolor=BG_CARD, alpha=0.8,
                 edgecolor="#374151")
    ax1.text(0.98, 0.05, info, transform=ax1.transAxes, fontsize=7,
             ha="right", va="bottom", bbox=props, color=TEXT_MUTED)

    # ── Panel 2: Seasonal comparison (normalised shape) ───────────────
    month = pd.Timestamp(tomorrow_str).month

    # Historical average for same month
    raw_monthly = raw_df[raw_df.index.month == month].copy()
    raw_monthly["hour"] = raw_monthly.index.hour
    seasonal_avg = raw_monthly.groupby("hour")["da_price_eur_mwh"].mean()

    # Normalise both to index 100 = daily mean
    today_mean = np.mean(preds) if np.mean(preds) != 0 else 1.0
    seasonal_mean = seasonal_avg.mean() if seasonal_avg.mean() != 0 else 1.0

    today_norm = (preds / today_mean) * 100
    seasonal_norm = (seasonal_avg.values / seasonal_mean) * 100

    ax2.plot(hours, today_norm, color=BLUE, linewidth=1.2, marker="o",
             markersize=3, label="Today's forecast")
    if len(seasonal_norm) == 24:
        ax2.plot(range(24), seasonal_norm, color=ORANGE, linewidth=1.2,
                 linestyle="--", marker="s", markersize=2, label="Seasonal average")

    ax2.axhline(100, color="#374151", linewidth=0.5)
    ax2.set_title("Shape vs seasonal average", fontsize=10)
    ax2.set_xlabel("Hour (UTC)")
    ax2.set_ylabel("Normalised price (index, avg=100)")
    ax2.set_xticks(range(0, 24, 3))
    ax2.legend(fontsize=7, framealpha=0.3)

    fig.tight_layout()
    fig.savefig(out_dir / "fig_forecast_tomorrow.png", dpi=PLOT_DPI,
                bbox_inches="tight", facecolor=BG_DARK)
    plt.close(fig)
    logger.info("Saved fig_forecast_tomorrow.png")


# =====================================================================
# Figure Generation Entry Point
# =====================================================================
def generate_figures(target_date: str, data_dir: Path, outputs_dir: Path) -> dict:
    """Generate both report figures and save to outputs/reports/{date}/."""
    out_dir = outputs_dir / "reports" / target_date
    out_dir.mkdir(parents=True, exist_ok=True)

    # Load data
    signal_df = pd.read_csv(outputs_dir / "curve_translation" / "signal_table.csv")
    signal_df["delivery_date"] = pd.to_datetime(signal_df["delivery_date"])

    delivery_df = pd.read_csv(outputs_dir / "curve_translation" / "delivery_periods.csv")
    delivery_df["delivery_date"] = pd.to_datetime(delivery_df["delivery_date"])

    oos_df = pd.read_parquet(data_dir / "processed" / "oos_predictions.parquet")
    raw_df = pd.read_parquet(data_dir / "processed" / "de_power_dataset.parquet")

    _generate_fig1(signal_df, delivery_df, target_date, out_dir)
    _generate_fig2(oos_df, delivery_df, raw_df, target_date, out_dir)

    return {
        "fig1": out_dir / "fig_signal_dashboard.png",
        "fig2": out_dir / "fig_forecast_tomorrow.png",
    }


# =====================================================================
# Raw Data Excel
# =====================================================================
def generate_raw_data_excel(target_date: str, outputs_dir: Path) -> Path:
    """
    Generate the companion Excel workbook with 3 trader-focused sheets:
      1. Hourly Forecast  — tomorrow's 24-hour price curve (the one thing
         traders actually paste into their own models)
      2. Signal History    — last 14 days of signals, premiums, confidence,
         invalidation — enough to see the trend, not overwhelming
      3. Fundamentals      — last 14 days of wind, gas, residual load, solar,
         spark spread — the drivers behind the signal
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    out_dir = outputs_dir / "reports" / target_date
    out_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = out_dir / "raw_data.xlsx"

    td = pd.Timestamp(target_date)

    # Styling
    header_fill = PatternFill(start_color="1a1f2e", end_color="1a1f2e", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=9)
    alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    today_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    thin_border = Border(bottom=Side(style="thin", color="E5E7EB"))
    num_2dp = '#,##0.00'
    num_1dp = '#,##0.0'

    def _write_headers(ws, headers, widths=None):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        if widths:
            for c, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(c)].width = w

    def _safe_float(val, decimals=2):
        if pd.isna(val) or val is None:
            return None
        return round(float(val), decimals)

    wb = Workbook()

    # ── Sheet 1: Hourly Forecast ──────────────────────────────────────
    # This is the sheet traders actually copy into their own spreadsheets.
    ws1 = wb.active
    ws1.title = "Hourly Forecast"

    oos_df = pd.read_parquet(PROJECT_ROOT / "data" / "processed" / "oos_predictions.parquet")
    tomorrow = td + pd.Timedelta(days=1)
    tmrw_tz = pd.Timestamp(str(tomorrow.date()), tz="UTC")
    tmrw_data = oos_df.loc[
        (oos_df.index >= tmrw_tz)
        & (oos_df.index < tmrw_tz + pd.Timedelta(days=1))
    ]
    if len(tmrw_data) == 0:
        tmrw_tz = pd.Timestamp(target_date, tz="UTC")
        tmrw_data = oos_df.loc[
            (oos_df.index >= tmrw_tz)
            & (oos_df.index < tmrw_tz + pd.Timedelta(days=1))
        ]

    headers = ["Hour (UTC)", "Forecast (€/MWh)", "Actual (€/MWh)", "Error (€/MWh)", "Block"]
    _write_headers(ws1, headers, widths=[12, 18, 16, 14, 10])

    peak_hours = set(range(8, 20))
    for r_idx, (ts, row) in enumerate(tmrw_data.iterrows(), 2):
        h = ts.hour
        pred = _safe_float(row["y_pred"])
        actual = _safe_float(row.get("y_actual"))
        error = _safe_float(pred - actual) if pred is not None and actual is not None else None
        block = "Peak" if h in peak_hours else "Offpeak"

        ws1.cell(row=r_idx, column=1, value=f"{h:02d}:00")
        c_pred = ws1.cell(row=r_idx, column=2, value=pred)
        c_pred.number_format = num_2dp
        c_act = ws1.cell(row=r_idx, column=3, value=actual)
        c_act.number_format = num_2dp
        c_err = ws1.cell(row=r_idx, column=4, value=error)
        c_err.number_format = num_2dp
        ws1.cell(row=r_idx, column=5, value=block)

        if h in peak_hours:
            for c in range(1, 6):
                ws1.cell(row=r_idx, column=c).fill = today_fill
        elif r_idx % 2 == 0:
            for c in range(1, 6):
                ws1.cell(row=r_idx, column=c).fill = alt_fill

    # Summary rows at the bottom
    sum_row = len(tmrw_data) + 3
    if len(tmrw_data) > 0:
        preds = tmrw_data["y_pred"]
        peak_preds = tmrw_data[tmrw_data.index.hour.isin(peak_hours)]["y_pred"]
        offpeak_preds = tmrw_data[~tmrw_data.index.hour.isin(peak_hours)]["y_pred"]

        for label, val, row_off in [
            ("Base avg", _safe_float(preds.mean()), 0),
            ("Peak avg", _safe_float(peak_preds.mean()), 1),
            ("Offpeak avg", _safe_float(offpeak_preds.mean()), 2),
            ("Peak premium", _safe_float(peak_preds.mean() - preds.mean()), 3),
            ("Min", _safe_float(preds.min()), 4),
            ("Max", _safe_float(preds.max()), 5),
            ("Spread (max-min)", _safe_float(preds.max() - preds.min()), 6),
        ]:
            r = sum_row + row_off
            ws1.cell(row=r, column=1, value=label).font = Font(bold=True)
            c = ws1.cell(row=r, column=2, value=val)
            c.number_format = num_2dp
            c.font = Font(bold=True)

    # ── Sheet 2: Signal History (last 14 days) ────────────────────────
    # Clean view of just the columns that matter for position decisions.
    ws2 = wb.create_sheet("Signal History")

    signal_df = pd.read_csv(outputs_dir / "curve_translation" / "signal_table.csv")
    signal_df["delivery_date"] = pd.to_datetime(signal_df["delivery_date"])
    sig_14 = signal_df[signal_df["delivery_date"] >= td - pd.Timedelta(days=13)].copy()

    # Select only the columns a trader needs
    keep_cols = [
        ("delivery_date", "Date", 12),
        ("signal_label_month_base", "Signal", 14),
        ("signal_month_base", "Premium (€)", 13),
        ("fv_month_base", "Fair Value (€)", 14),
        ("curve_proxy_30d", "Curve Proxy (€)", 15),
        ("month_confidence", "Confidence", 12),
        ("month_position_size", "Position Size", 13),
        ("spark_spread_actual", "Spark Spread (€)", 16),
        ("any_invalidation", "Invalidated", 12),
    ]
    keep_cols = [(src, label, w) for src, label, w in keep_cols if src in sig_14.columns]

    _write_headers(ws2, [label for _, label, _ in keep_cols],
                   widths=[w for _, _, w in keep_cols])

    for r_idx, (_, row) in enumerate(sig_14.iterrows(), 2):
        is_today = (row["delivery_date"] == td)
        for c_idx, (src, _, _) in enumerate(keep_cols, 1):
            val = row[src]
            if isinstance(val, pd.Timestamp):
                val = str(val.date())
            elif isinstance(val, (np.floating, float)):
                val = _safe_float(val)
            elif isinstance(val, (np.bool_, bool)):
                val = "YES" if val else ""
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            if isinstance(val, (int, float)) and val is not None:
                cell.number_format = num_2dp
            if is_today:
                cell.fill = today_fill
            elif r_idx % 2 == 0:
                cell.fill = alt_fill
            cell.border = thin_border

    # ── Sheet 3: Fundamentals (last 14 days) ──────────────────────────
    # The market drivers — what moved and why.
    ws3 = wb.create_sheet("Fundamentals")

    del_df = pd.read_csv(outputs_dir / "curve_translation" / "delivery_periods.csv")
    del_df["delivery_date"] = pd.to_datetime(del_df["delivery_date"])
    del_14 = del_df[del_df["delivery_date"] >= td - pd.Timedelta(days=13)]

    fund_cols = [
        ("delivery_date", "Date", 12),
        ("base_pred", "Base Fcst (€)", 13),
        ("base_actual", "Base Actual (€)", 14),
        ("peak_pred", "Peak Fcst (€)", 13),
        ("avg_wind_mw", "Wind (MW)", 12),
        ("avg_solar_mw", "Solar (MW)", 12),
        ("avg_residual_load_mw", "Res. Load (MW)", 14),
        ("gas_price", "Gas (€/MWh)", 12),
        ("avg_res_penetration", "RES %", 8),
    ]
    fund_cols = [(src, label, w) for src, label, w in fund_cols if src in del_14.columns]

    _write_headers(ws3, [label for _, label, _ in fund_cols],
                   widths=[w for _, _, w in fund_cols])

    for r_idx, (_, row) in enumerate(del_14.iterrows(), 2):
        is_today = (row["delivery_date"] == td)
        for c_idx, (src, _, _) in enumerate(fund_cols, 1):
            val = row[src]
            if isinstance(val, pd.Timestamp):
                val = str(val.date())
            elif isinstance(val, (np.floating, float)):
                val = _safe_float(val)
            cell = ws3.cell(row=r_idx, column=c_idx, value=val)
            if isinstance(val, (int, float)) and val is not None:
                cell.number_format = num_1dp
            if is_today:
                cell.fill = today_fill
            elif r_idx % 2 == 0:
                cell.fill = alt_fill
            cell.border = thin_border

    ws3.freeze_panes = "A2"

    # ── Save ──────────────────────────────────────────────────────────
    wb.save(str(xlsx_path))
    logger.info("Saved raw_data.xlsx")
    return xlsx_path


# =====================================================================
# PDF Report
# =====================================================================
def render_pdf(
    target_date: str,
    outputs_dir: Path,
    figure_paths: dict,
    excel_path: Path,
    config: dict,
) -> Path:
    """Render the 3-page PDF report using Jinja2 + WeasyPrint."""
    from jinja2 import Environment, FileSystemLoader
    from weasyprint import HTML

    out_dir = outputs_dir / "reports" / target_date
    out_dir.mkdir(parents=True, exist_ok=True)
    pdf_path = out_dir / "daily_report.pdf"

    # Load data for template
    signal_df = pd.read_csv(outputs_dir / "curve_translation" / "signal_table.csv")
    signal_df["delivery_date"] = pd.to_datetime(signal_df["delivery_date"])
    td = pd.Timestamp(target_date)
    today_row = signal_df[signal_df["delivery_date"] == td]

    if len(today_row) == 0:
        today_row = signal_df.tail(1)

    r = today_row.iloc[0]

    delivery_df = pd.read_csv(outputs_dir / "curve_translation" / "delivery_periods.csv")
    delivery_df["delivery_date"] = pd.to_datetime(delivery_df["delivery_date"])
    del_row = delivery_df[delivery_df["delivery_date"] == td]
    d = del_row.iloc[0] if len(del_row) > 0 else pd.Series()

    # AI briefing — extract narrative sections, keep citations for grounding
    briefing_path = outputs_dir / "ai_intelligence" / target_date / "briefing.txt"
    ai_sections = {}  # section_name -> text
    if briefing_path.exists():
        full_text = briefing_path.read_text(encoding="utf-8")
        lines = full_text.strip().splitlines()
        current_section = None
        for line in lines:
            stripped = line.strip()
            if stripped in ("SIGNAL SUMMARY", "FUNDAMENTAL DRIVERS",
                            "TOP MODEL FEATURES", "INTRADAY SHAPE",
                            "INVALIDATION ASSESSMENT"):
                current_section = stripped
                ai_sections[current_section] = []
            elif current_section and stripped and not stripped.startswith("===") and not stripped.startswith("────") and not stripped.startswith("DE_LU"):
                if stripped:
                    ai_sections[current_section].append(stripped)
        # Join each section — full text, no truncation
        for k in ai_sections:
            ai_sections[k] = " ".join(ai_sections[k])

    # ── Helper ────────────────────────────────────────────────────────
    def _fv(val, fmt=".2f"):
        if pd.isna(val) or val is None:
            return None
        return f"{val:{fmt}}"

    # ── Signal: use current, fall back to last available ─────────────
    sig_label_raw = str(r.get("signal_label_month_base", "N/A"))
    has_signal = sig_label_raw in (
        "STRONG_BUY", "BUY", "HOLD", "SELL", "STRONG_SELL", "INVALIDATED"
    )

    # If today has no signal, find the last row that does
    prior_sig = None
    prior_sig_date = None
    if not has_signal:
        valid_signals = signal_df[
            signal_df["signal_label_month_base"].isin(
                ["STRONG_BUY", "BUY", "HOLD", "SELL", "STRONG_SELL", "INVALIDATED"]
            ) & (signal_df["delivery_date"] < td)
        ]
        if len(valid_signals) > 0:
            prior_sig = valid_signals.iloc[-1]
            prior_sig_date = str(prior_sig["delivery_date"].date())

    sig_label = sig_label_raw if has_signal else (
        str(prior_sig["signal_label_month_base"]) if prior_sig is not None else "HOLD"
    )
    sig_class = {
        "STRONG_BUY": "sig-strong-buy",
        "BUY": "sig-buy",
        "HOLD": "sig-hold",
        "SELL": "sig-sell",
        "STRONG_SELL": "sig-strong-sell",
        "INVALIDATED": "sig-invalidated",
    }.get(sig_label, "sig-hold")

    # Pull signal metrics from current row or prior
    sig_src = r if has_signal else (prior_sig if prior_sig is not None else r)

    # ── Invalidation flags ───────────────────────────────────────────
    inv_flags = [
        ("Gas spike", r.get("inv_gas_spike", False)),
        ("Wind revision", r.get("inv_wind_revision", False)),
        ("Residual load swing", r.get("inv_residual_load_swing", False)),
        ("Negative price regime", r.get("inv_negative_price_regime", False)),
        ("High model error", r.get("inv_high_error_regime", False)),
    ]
    inv_active = sum(1 for _, v in inv_flags if v)

    # ── Trailing 7-day averages ──────────────────────────────────────
    trailing_7 = delivery_df[
        (delivery_df["delivery_date"] < td)
        & (delivery_df["delivery_date"] >= td - pd.Timedelta(days=7))
    ]

    def _trail(col, fmt=".1f"):
        if col in trailing_7.columns and len(trailing_7) > 0:
            val = trailing_7[col].mean()
            if pd.notna(val):
                return f"{val:{fmt}}"
        return None

    # ── Peak-base premium ────────────────────────────────────────────
    peak_prem = None
    if len(d) > 0 and pd.notna(d.get("peak_pred")) and pd.notna(d.get("base_pred")):
        peak_prem = _fv(d["peak_pred"] - d["base_pred"], "+.1f")

    # ── Forecast accuracy (if actuals exist) ─────────────────────────
    forecast_error = None
    base_actual = None
    if len(d) > 0 and pd.notna(d.get("base_actual")) and pd.notna(d.get("base_pred")):
        base_actual = _fv(d["base_actual"], ".1f")
        err = d["base_actual"] - d["base_pred"]
        forecast_error = _fv(err, "+.1f")

    # ── Prediction intervals from delivery_periods ───────────────────
    p10 = _fv(d.get("base_pred_p10") if len(d) > 0 else None, ".1f")
    p90 = _fv(d.get("base_pred_p90") if len(d) > 0 else None, ".1f")

    # ── Offpeak forecast ─────────────────────────────────────────────
    offpeak_pred = _fv(d.get("offpeak_pred") if len(d) > 0 else None, ".1f")

    # ── Intraday shape metrics ───────────────────────────────────────
    intraday_spread = _fv(d.get("intraday_spread") if len(d) > 0 else None, ".1f")
    neg_hours = None
    if "neg_hours_pred" in r.index and pd.notna(r.get("neg_hours_pred")):
        neg_hours = int(r["neg_hours_pred"])

    # ── Regime ───────────────────────────────────────────────────────
    is_thermal = r.get("is_thermal_regime", False)
    regime_label = "Thermal" if is_thermal else "RES-dominated"

    # ── Curve proxies ────────────────────────────────────────────────
    curve_7d = _fv(r.get("curve_proxy_7d"), ".1f")
    curve_30d = _fv(r.get("curve_proxy_30d"), ".1f")
    curve_90d = _fv(r.get("curve_proxy_90d"), ".1f")

    # ── RES penetration ──────────────────────────────────────────────
    res_pct = _fv(r.get("avg_res_penetration"), ".0%") if pd.notna(r.get("avg_res_penetration")) else (
        _fv(d.get("avg_res_penetration") if len(d) > 0 else None, ".0%")
    )

    # ── Model performance (static, from report) ─────────────────────
    model_mae = None
    model_skill = None
    model_dir_acc = None
    report_path = outputs_dir / "model_performance_report.txt"
    if report_path.exists():
        for line in report_path.read_text(encoding="utf-8").splitlines():
            s = line.strip()
            if s.startswith("Hourly") and "MAE:" in s:
                model_mae = s.split("MAE:")[1].strip().split()[0]
            elif s.startswith("Directional"):
                model_dir_acc = s.split(":")[1].strip().split()[0]
            elif s.startswith("Skill"):
                model_skill = s.split(":")[1].strip().split()[0]

    # ── Backtest stats (from curve_translation_report) ───────────────
    backtest_sharpe = None
    backtest_hit_rate = None
    bt_path = outputs_dir / "curve_translation" / "curve_translation_report.txt"
    if bt_path.exists():
        for line in bt_path.read_text(encoding="utf-8").splitlines():
            s = line.strip()
            if "Sharpe" in s and ":" in s:
                try:
                    backtest_sharpe = s.split(":")[1].strip().split()[0]
                except (IndexError, ValueError):
                    pass
            elif "Hit rate" in s and ":" in s:
                try:
                    backtest_hit_rate = s.split(":")[1].strip().split()[0]
                except (IndexError, ValueError):
                    pass

    template_data = {
        "target_date": target_date,
        "generated_time": datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"),
        # Signal
        "sig_label": sig_label,
        "sig_class": sig_class,
        "sig_label_display": sig_label.replace("_", " "),
        "is_prior_signal": not has_signal and prior_sig is not None,
        "prior_sig_date": prior_sig_date,
        # Signal metrics (from current or fallback to prior)
        "fv_month_base": _fv(sig_src.get("fv_month_base"), ".1f"),
        "curve_proxy_30d": curve_30d,
        "signal_month_base": _fv(sig_src.get("signal_month_base"), "+.1f") if pd.notna(sig_src.get("signal_month_base")) else None,
        "month_confidence": _fv(sig_src.get("month_confidence"), ".2f"),
        "month_position_size": _fv(sig_src.get("month_position_size"), ".1f"),
        "spark_spread_actual": _fv(r.get("spark_spread_actual"), ".1f"),
        # Curve proxies
        "curve_7d": curve_7d,
        "curve_90d": curve_90d,
        # Tomorrow's price
        "base_pred": _fv(d.get("base_pred") if len(d) > 0 else None, ".1f"),
        "peak_pred": _fv(d.get("peak_pred") if len(d) > 0 else None, ".1f"),
        "offpeak_pred": offpeak_pred,
        "peak_premium": peak_prem,
        "p10": p10,
        "p90": p90,
        # Forecast accuracy
        "base_actual": base_actual,
        "forecast_error": forecast_error,
        # Market context
        "avg_residual_load_mw": _fv(d.get("avg_residual_load_mw") if len(d) > 0 else None, ",.0f"),
        "avg_wind_mw": _fv(d.get("avg_wind_mw") if len(d) > 0 else None, ",.0f"),
        "avg_solar_mw": _fv(d.get("avg_solar_mw") if len(d) > 0 else None, ",.0f"),
        "gas_price": _fv(d.get("gas_price") if len(d) > 0 else None, ".1f"),
        "trail_wind": _trail("avg_wind_mw", ",.0f"),
        "trail_gas": _trail("gas_price", ".1f"),
        "trail_res_load": _trail("avg_residual_load_mw", ",.0f"),
        "trail_base": _trail("base_actual", ".1f"),
        # Regime & shape
        "regime_label": regime_label,
        "res_pct": res_pct,
        "intraday_spread": intraday_spread,
        "neg_hours": neg_hours,
        # Model credibility
        "model_mae": model_mae,
        "model_dir_acc": model_dir_acc,
        "backtest_sharpe": backtest_sharpe,
        "backtest_hit_rate": backtest_hit_rate,
        # Invalidation
        "inv_flags": inv_flags,
        "any_invalidation": bool(r.get("any_invalidation", False)),
        "inv_active": inv_active,
        # AI briefing sections
        "ai_signal": ai_sections.get("SIGNAL SUMMARY", ""),
        "ai_fundamentals": ai_sections.get("FUNDAMENTAL DRIVERS", ""),
        "ai_shape": ai_sections.get("INTRADAY SHAPE", ""),
        # Figures
        "fig1_path": str(figure_paths.get("fig1", "")),
        "fig2_path": str(figure_paths.get("fig2", "")),
    }

    # Render template
    env = Environment(loader=FileSystemLoader(str(TEMPLATES_DIR)))
    template = env.get_template("report.html")
    rendered_html = template.render(**template_data)

    # Convert to PDF
    HTML(string=rendered_html, base_url=str(out_dir)).write_pdf(str(pdf_path))
    logger.info("Saved daily_report.pdf")
    return pdf_path


# =====================================================================
# Master Orchestrator
# =====================================================================
def run_report_generation(target_date: str = None) -> dict:
    """Generate figures, Excel, and PDF for the given date."""
    _setup_logging()
    logger.info("=" * 60)
    logger.info("Report Generation Pipeline")
    logger.info("=" * 60)

    if target_date is None:
        # Default to today; if today isn't in the data, use the latest date
        # that has a real signal (skip NO_DATA terminal rows)
        today_str = str(datetime.utcnow().date())
        signal_df = pd.read_csv(OUTPUTS_DIR / "curve_translation" / "signal_table.csv")
        signal_df["delivery_date"] = pd.to_datetime(signal_df["delivery_date"])
        available_dates = set(signal_df["delivery_date"].dt.strftime("%Y-%m-%d"))

        if today_str in available_dates:
            target_date = today_str
        else:
            # Pick the last date that has an actual signal (not NO_DATA)
            valid = signal_df[signal_df["signal_label_month_base"].isin(
                ["STRONG_BUY", "BUY", "HOLD", "SELL", "STRONG_SELL", "INVALIDATED"]
            )]
            if len(valid) > 0:
                target_date = str(valid["delivery_date"].max().date())
            else:
                target_date = str(signal_df["delivery_date"].max().date())
    logger.info("Target date: %s", target_date)

    out_dir = REPORTS_DIR / target_date
    out_dir.mkdir(parents=True, exist_ok=True)

    # Load config
    config_path = PROJECT_ROOT / "server_config.json"
    config = {}
    if config_path.exists():
        config = json.loads(config_path.read_text())

    # Generate figures
    logger.info("Generating figures …")
    figure_paths = generate_figures(target_date, DATA_DIR, OUTPUTS_DIR)

    # Generate Excel
    logger.info("Generating raw data Excel …")
    excel_path = generate_raw_data_excel(target_date, OUTPUTS_DIR)

    # Generate PDF
    logger.info("Generating PDF report …")
    pdf_path = render_pdf(target_date, OUTPUTS_DIR, figure_paths, excel_path, config)

    logger.info("Report generation complete: %s", out_dir)

    return {
        "pdf": pdf_path,
        "excel": excel_path,
        "fig1": figure_paths.get("fig1"),
        "fig2": figure_paths.get("fig2"),
        "target_date": target_date,
    }
